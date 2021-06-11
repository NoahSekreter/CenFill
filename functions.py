import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from shutil import copy2
from datetime import datetime
import os
import warnings
from collections import defaultdict
import xlrd
import csv
import requests
# import json
# import urllib.request
# import pyexcel as p


# Format the date into month, day, year. If unsuccessful, return the cell unchanged
def format_date(matrix, row, col):
    try:
        return datetime.strptime(matrix[row][col], '%Y-%m-%d').strftime('%m/%d/%Y')
    except(ValueError, Exception):
        return matrix[row][col]


# Get age from a birthday
def get_age(birthday_string):
    try:
        birthday = datetime.strptime(str(birthday_string), '%m/%d/%Y')
        today = datetime.today()
        return today.year - birthday.year - ((today.month, today.day) < (birthday.month, birthday.day))
    except(ValueError, Exception):
        return 0


# Return current year
def get_current_year():
    return datetime.today().year


# Check which file has been inputted
def check_input_file(file):
    wb = load_workbook(file, read_only=True)
    ws = wb.active

    # ICHRA File
    if "Age" in str(ws.cell(row=1, column=5).value):
        return 3
    # 1095 File
    elif "LAST NAME" in str(ws.cell(row=7, column=1).value):
        return 2
    # Census file
    elif "First Name" in str(ws.cell(row=1, column=1).value):
        return 1
    else:
        return 0


# Rename the file correctly
def rename(file_loc, old_name, template):
    # Get base name, stripping away file location
    company_name = str(os.path.basename(file_loc).split('Census Filler', 1)[0])
    carrier_dict = {"1": "Anthem MEWA Census ",
                    "2": "MMO Census ",
                    "3": "UHC GRX Census ",
                    "4": "Cigna GRX Census ",
                    "5": "SummaCare Census ",
                    "6": "Aetna 51-99 Census ",
                    "7": "Aetna CAT 100+ Census ",
                    "8": "Aetna Tier 100+ Census ",
                    "9": "Anthem ACA Census "}
    company_name += carrier_dict[template] + datetime.now().strftime("%m.%d.%Y") + ".xlsx"
    new_name = file_loc.replace(os.path.basename(file_loc), '') + company_name  # Get target directory
    if os.path.exists(new_name):
        os.remove(new_name)
    os.rename(old_name, new_name)
    # TODO: This is bugged. It doesn't change the Anthem census back to .xls
    # if template is "1":
    #     p.save_book_as(file_name=new_name, dest_file_name=new_name[:-1])
    #     os.remove(new_name)


def basename(raw_name):
    return str(os.path.basename(raw_name)).strip(".xlsx")


# Convert an xlsx file to a cvs file
def xlsx_to_csv(input_xlsx):
    wb = xlrd.open_workbook(input_xlsx)
    sh = wb.sheet_by_index(0)

    # Create copy of cvs template and place in target location
    base = input_xlsx.replace(os.path.basename(input_xlsx), '')  # Get target directory
    copy2('templates/Completed FT Census.csv', base)
    new_loc = os.path.join(base, os.path.basename('templates/Completed FT Census.csv'))

    csv_file = open(new_loc, 'w', encoding='utf-8', newline='')
    wr = csv.writer(csv_file, quoting=csv.QUOTE_ALL)

    for rownum in range(0, int(sh.nrows)):
        wr.writerow(sh.row_values(rownum))

    csv_file.close()
    os.remove(input_xlsx)


# Create a matrix from an enrollment sheet to more easily copy info onto an excel spreadsheet
def generate_client_matrix(client, sheet_id):
    # Prepare a sub-function to assign the columns in the correct matrix position
    def transfer_column(input_col, matrix_col):
        # Go through each row
        for input_row in range(len(current_sheet.rows)):
            client_matrix[input_row][matrix_col] = current_sheet.rows[input_row].cells[input_col].value

    # Get the current sheet and prepare the client matrix
    current_sheet = client.Sheets.get_sheet(sheet_id)
    client_matrix = [["" for x in range(8)] for y in current_sheet.rows]

    # Locate every relevant column and transfer the column's data to the correct position within the matrix
    for col in range(len(current_sheet.columns)):
        title = str(current_sheet.columns[col].title).lower()
        if "first name" in title:
            transfer_column(col, 0)
        elif "last name" in title:
            transfer_column(col, 1)
        elif "date of birth" in title:
            transfer_column(col, 2)
        elif "gender" in title:
            transfer_column(col, 3)
        elif "zip" in title:
            transfer_column(col, 4)
        elif "relation" in title:
            transfer_column(col, 5)
        elif "termination" in title:
            transfer_column(col, 6)
        elif "waiver" in title:
            transfer_column(col, 7)

    # Remove any termed/waived/empty members from the matrix

    return client_matrix


# Create a census filler from a enrollment sheet
def generate_spreadsheet(file_loc, client_matrix, name):
    copy2('templates/Census Filler.xlsx', file_loc)  # Copy the new file into the target directory
    new_dir = os.path.join(file_loc, 'Census Filler.xlsx')  # Get absolute path of the copied file
    workbook = load_workbook(new_dir)  # Load the copied file
    wb = workbook.active
    n = 2

    # Create dictionaries to make copying data easier
    rel_dict = {"Subscriber": "E",
                "Spouse": "S",
                "Child": "D",
                "COBRA": "C",
                defaultdict: ""}
    rel_num_dict = {"E": 0,
                    "S": 1,
                    "D": 2,
                    defaultdict: ""}
    med_tier_dict = {"E": "EE",
                     "ES": "ES",
                     "EC": "EC",
                     "F": "F",
                     "C": "E",
                     defaultdict: ""}

    # Copy all info from the matrix onto the census filler
    term_dep_count = 1
    for row in range(3, len(client_matrix)):
        # Skip if the current member is termed
        if client_matrix[row][6] is not None:
            continue
        # Skip if the current member is waived
        if client_matrix[row][7] is not None:
            if "Waiver" in client_matrix[row][7]:
                continue
        # Skip if first and last name are not present
        if client_matrix[row][0] is None and client_matrix[row][1] is None:
            continue

        # Check if a dependant is part of a termed subscriber, if so then ignore them
        if "Spouse" in client_matrix[row][5] or "Child" in client_matrix[row][5]:
            term_dep_count += 1
            if client_matrix[row - term_dep_count][6] is not None:
                continue
        else:
            term_dep_count = 1

        # Copy the required information
        wb.cell(row=n, column=1).value = client_matrix[row][0]  # First name
        wb.cell(row=n, column=2).value = client_matrix[row][1]  # Last name
        wb.cell(row=n, column=3).value = format_date(client_matrix, row, 2)  # DoB
        wb.cell(row=n, column=4).value = client_matrix[row][3]  # Gender
        wb.cell(row=n, column=5).value = client_matrix[row][4]  # Zip Code
        wb.cell(row=n, column=8).value = rel_dict[client_matrix[row][5]]  # Relationship code
        wb.cell(row=n, column=9).value = rel_num_dict[wb.cell(row=n, column=8).value]  # Relationship 1
        wb.cell(row=n, column=10).value = 1 if wb.cell(row=n, column=9).value is 0 else 0  # Relationship 2
        try:
            wb.cell(row=n, column=12).value = datetime.strptime(str(wb.cell(row=n, column=3).value),
                                                                '%m/%d/%Y').strftime('%m/%y')  # DoB again
        except(ValueError, Exception):
            wb.cell(row=n, column=12).value = None
        n += 1

    # Go through the census again to generate medical tiers and child count
    child_count = 0
    for row in range(2, n):
        if wb.cell(row=row, column=8).value is "E":
            child_count = 0
            if wb.cell(row=row + 1, column=8).value is "S":
                wb.cell(row=row, column=6).value = "F" if wb.cell(row=row + 2, column=8).value is "D" else "ES"
            elif wb.cell(row=row + 1, column=8).value is "D":
                wb.cell(row=row, column=6).value = "EC"
            else:
                wb.cell(row=row, column=6).value = "E"
        elif wb.cell(row=row, column=8).value is "S" or "D":
            wb.cell(row=row, column=6).value = wb.cell(row=row - 1, column=6).value
            if wb.cell(row=row, column=8).value is "D":
                child_count += 1
                wb.cell(row=row - child_count - (1 if wb.cell(row=row - child_count, column=8).value is "S" else 0),
                        column=11).value = child_count
        wb.cell(row=row, column=7).value = med_tier_dict[wb.cell(row=row, column=6).value]
        if get_age(wb.cell(row=row, column=3).value) >= 65:
            for col in range(1, 13):
                wb.cell(row=row, column=col).fill = PatternFill(fill_type='solid', start_color='B8FAFA',
                                                                end_color='B8FAFA')
        elif get_age(wb.cell(row=row, column=3).value) >= 26 and wb.cell(row=row, column=8).value is "D":
            for col in range(1, 13):
                wb.cell(row=row, column=col).fill = PatternFill(fill_type='solid', start_color='F2F277',
                                                                end_color='F2F277')

    workbook.save(filename=new_dir)

    # Rename the spreadsheet
    company_name = name + " Census Filler " + datetime.now().strftime("%m.%d.%Y") + ".xlsx"
    new_name = file_loc + '/' + company_name  # Get target directory
    if os.path.exists(new_name):
        os.remove(new_name)
    os.rename(new_dir, new_name)
    return new_name


# Moves info over onto new template
def generate_carrier_spreadsheets(file_loc, template):
    warnings.filterwarnings("ignore")
    i_file = load_workbook(filename=file_loc)
    i_wb = i_file.active
    i_n = 2
    o_n = 0
    ee_id = 0
    mmo_dep = 15
    o_dir, o_wb, o_file, new_dir = None, None, None, None
    health_conv = {'EE': 'S',
                   'ES': 'T',
                   'EC': 'C',
                   'F': 'F',
                   defaultdict: ''}
    member_type = {'E': 'Employee',
                   'S': 'Spouse',
                   'D': 'Child',
                   defaultdict: ''}
    tier_type = {'E': 'Employee',
                 'ES': 'Employee + Spouse',
                 'EC': 'Employee + Child',
                 'F': 'Family',
                 'W': 'Waiver',
                 defaultdict: ''}
    contract_type = {'E': 'Single',
                     'ES': 'EE + SP',
                     'EC': 'EE + CH',
                     'F': 'Family',
                     'W': 'Waive',
                     defaultdict: ''}
    template_dict = {'1': 'templates/Anthem MEWA Quoting Census.xlsx',
                     '2': 'templates/MMO Census.xlsx',
                     '3': 'templates/UHC GRX Quote Coversheet & Census.xlsx',
                     '4': 'templates/Cigna GRX Census.xlsx',
                     '5': 'templates/SummaCare Census.xlsx',
                     '6': 'templates/51-99 Aetna Census.xlsx',
                     '7': 'templates/Aetna CAT 100+ Census.xlsx',
                     '8': 'templates/Aetna Tier 100+ Census.xlsx',
                     '9': 'templates/Anthem ACA Census.xlsx',
                     defaultdict: None}
    o_dir = template_dict[template]
    if o_dir is not None:
        new_loc = file_loc.replace(os.path.basename(file_loc), '')  # Get target directory
        copy2(o_dir, new_loc)  # Copy the new file into the target directory
        new_dir = os.path.join(new_loc, os.path.basename(o_dir))  # Get absolute path of the copied file
        o_file = load_workbook(new_dir)  # Load the copied file
        o_wb = o_file.active

    # Go through the input census and reformat for whatever the template is
    while o_file is not None and i_wb.cell(row=i_n, column=1).value is not None:
        # Is this an employee?
        if i_wb.cell(row=i_n, column=8).value is 'E':
            if template is '1':  # Anthem
                o_n += 1
                if i_wb.cell(row=i_n, column=3).value is not None:
                    try:
                        o_wb.cell(row=o_n + 6, column=2).value = datetime.strptime(
                            str(i_wb.cell(row=i_n, column=3).value), '%m/%d/%Y').strftime('%m/%y')  # Birthday
                    except(ValueError, Exception):
                        o_wb.cell(row=o_n + 6, column=2).value = datetime.strptime(
                            str(i_wb.cell(row=i_n, column=3).value), '%Y-%m-%d 00:00:00').strftime('%m/%y')  # Birthday
                o_wb.cell(row=o_n + 6, column=4).value = i_wb.cell(row=i_n, column=4).value  # Gender
                o_wb.cell(row=o_n + 6, column=8).value = i_wb.cell(row=i_n, column=11).value  # Child Count
                o_wb.cell(row=o_n + 6, column=9).value = i_wb.cell(row=i_n, column=5).value  # Zip Code
                o_wb.cell(row=o_n + 6, column=10).value = 'A'  # Status
                o_wb.cell(row=o_n + 6, column=11).value = health_conv[i_wb.cell(row=i_n, column=7).value]  # Health Cov
            if template is '2':  # MMO
                o_n += 1
                mmo_dep = 15
                o_wb.cell(row=o_n + 7, column=1).value = i_wb.cell(row=i_n, column=2).value  # Last name
                try:
                    o_wb.cell(row=o_n + 7, column=2).value = datetime.strptime(
                        str(i_wb.cell(row=i_n, column=3).value), '%m/%d/%Y').strftime('%#m/%#d/%Y')  # Birthday
                except(ValueError, Exception):
                    o_wb.cell(row=o_n + 7, column=2).value = i_wb.cell(row=i_n, column=3).value
                o_wb.cell(row=o_n + 7, column=3).value = i_wb.cell(row=i_n, column=4).value  # Gender
                o_wb.cell(row=o_n + 7, column=4).value = 'N'
                o_wb.cell(row=o_n + 7, column=5).value = 'N'
                o_wb.cell(row=o_n + 7, column=12).value = 'Health'

        # Check for spouse
        if ('ES' in i_wb.cell(row=i_n, column=6).value or 'F' in i_wb.cell(row=i_n, column=6).value) and i_wb.cell(
                row=i_n, column=8).value is 'S':
            if template is '1':  # Anthem
                if i_wb.cell(row=i_n, column=3).value is not None:
                    try:
                        o_wb.cell(row=o_n + 6, column=5).value = datetime.strptime(
                            str(i_wb.cell(row=i_n, column=3).value), '%m/%d/%Y').strftime('%m/%y')  # Spouse Birthday
                    except(ValueError, Exception):
                        o_wb.cell(row=o_n + 6, column=5).value = datetime.strptime(
                            str(i_wb.cell(row=i_n, column=3).value), '%Y-%m-%d 00:00:00').strftime(
                            '%m/%y')  # Spouse Birthday
                o_wb.cell(row=o_n + 6, column=7).value = i_wb.cell(row=i_n, column=4).value  # Spouse Gender
            if template is '2':  # MMO
                if i_wb.cell(row=i_n, column=3).value is not None:
                    try:
                        o_wb.cell(row=o_n + 7, column=6).value = datetime.strptime(
                            str(i_wb.cell(row=i_n, column=3).value), '%m/%d/%Y').strftime(
                            '%#m/%#d/%Y')  # Spouse Birthday
                    except(ValueError, Exception):
                        o_wb.cell(row=o_n + 7, column=6).value = datetime.strptime(
                            str(i_wb.cell(row=i_n, column=3).value), '%Y-%m-%d 00:00:00').strftime(
                            '%#m/%#d/%Y')  # Spouse Birthday
                o_wb.cell(row=o_n + 7, column=7).value = i_wb.cell(row=i_n, column=4).value  # Spouse Gender
                o_wb.cell(row=o_n + 7, column=8).value = 'N'
                o_wb.cell(row=o_n + 7, column=9).value = 'N'

        # Check for children
        if ('EC' in i_wb.cell(row=i_n, column=6).value or 'F' in i_wb.cell(row=i_n, column=6).value) and i_wb.cell(
                row=i_n, column=8).value is 'D':
            # Anthem doesn't need child info
            if template is '2':  # MMO
                if i_wb.cell(row=i_n, column=3).value is not None:
                    try:
                        o_wb.cell(row=o_n + 7, column=mmo_dep).value = datetime.strptime(
                            str(i_wb.cell(row=i_n, column=3).value), '%m/%d/%Y').strftime(
                            '%#m/%#d/%Y')  # Child Birthday
                    except(ValueError, Exception):
                        o_wb.cell(row=o_n + 7, column=mmo_dep).value = datetime.strptime(
                            str(i_wb.cell(row=i_n, column=3).value), '%Y-%m-%d 00:00:00').strftime(
                            '%#m/%#d/%Y')  # Child Birthday
                o_wb.cell(row=o_n + 7, column=mmo_dep + 1).value = i_wb.cell(row=i_n, column=4).value  # Child Gender
                o_wb.cell(row=o_n + 7, column=mmo_dep + 2).value = 'N'
                o_wb.cell(row=o_n + 7, column=mmo_dep + 3).value = 'N'
                mmo_dep += 4

        # UHC is more or less a line-by-line copy
        if template is '3':  # UHC GRX
            o_n += 1
            o_wb.cell(row=o_n + 9, column=3).value = i_wb.cell(row=i_n, column=8).value  # Rel Code
            o_wb.cell(row=o_n + 9, column=4).value = i_wb.cell(row=i_n, column=2).value.lower().title()  # Last Name
            o_wb.cell(row=o_n + 9, column=5).value = i_wb.cell(row=i_n, column=1).value.lower().title()  # First Name
            o_wb.cell(row=o_n + 9, column=6).value = i_wb.cell(row=i_n, column=3).value  # Birthday
            o_wb.cell(row=o_n + 9, column=7).value = i_wb.cell(row=i_n, column=4).value  # Gender
            o_wb.cell(row=o_n + 9, column=8).value = i_wb.cell(row=i_n, column=5).value  # Zip
            if o_wb.cell(row=o_n + 9, column=3).value is 'E':
                ee_id += 1
                o_wb.cell(row=o_n + 9, column=9).value = i_wb.cell(row=i_n, column=6).value  # Medical Tier
                o_wb.cell(row=o_n + 9, column=10).value = 'Active'
            o_wb.cell(row=o_n + 9, column=2).value = ee_id

        # Very similar to the UHC census
        if template is '4':  # Cigna GRX
            o_n += 1
            o_wb.cell(row=o_n + 6, column=1).value = str(i_wb.cell(row=i_n, column=2).value).lower().title() if \
                i_wb.cell(row=i_n, column=2).value is not None else ''  # Last Name
            o_wb.cell(row=o_n + 6, column=2).value = str(i_wb.cell(row=i_n, column=1).value).lower().title() if \
                i_wb.cell(row=i_n, column=1).value is not None else ''  # First Name
            o_wb.cell(row=o_n + 6, column=3).value = i_wb.cell(row=i_n, column=5).value  # Zip
            o_wb.cell(row=o_n + 6, column=4).value = i_wb.cell(row=i_n, column=3).value  # Birthday
            o_wb.cell(row=o_n + 6, column=5).value = i_wb.cell(row=i_n, column=4).value  # Gender
            o_wb.cell(row=o_n + 6, column=6).value = i_wb.cell(row=i_n, column=10).value \
                if i_wb.cell(row=i_n, column=10).value != '' else '2'  # Relationship Code
            o_wb.cell(row=o_n + 6, column=7).value = i_wb.cell(row=i_n, column=6).value  # Medical Tier
            if o_wb.cell(row=o_n + 6, column=7).value is 'E':
                o_wb.cell(row=o_n + 6, column=7).value = 'EE'

        if template is '5':  # SummaCare
            o_n += 1
            o_wb.cell(row=o_n + 6, column=1).value = \
                i_wb.cell(row=i_n, column=1).value + ' ' + i_wb.cell(row=i_n, column=2).value  # Name
            o_wb.cell(row=o_n + 6, column=2).value = i_wb.cell(row=i_n, column=4).value  # Sex
            o_wb.cell(row=o_n + 6, column=3).value = get_age(i_wb.cell(row=i_n, column=3).value)  # Age
            o_wb.cell(row=o_n + 6, column=4).value = member_type[i_wb.cell(row=i_n, column=8).value]  # Member Type
            o_wb.cell(row=o_n + 6, column=5).value = contract_type[i_wb.cell(row=i_n, column=6).value]  # Contract Type

        if template is '6':  # Aetna 51-99
            o_n += 1
            o_wb.cell(row=o_n + 10, column=1).value = i_wb.cell(row=i_n, column=2).value  # Last Name
            o_wb.cell(row=o_n + 10, column=2).value = i_wb.cell(row=i_n, column=1).value  # First Name
            o_wb.cell(row=o_n + 10, column=4).value = i_wb.cell(row=i_n, column=5).value  # Zip
            o_wb.cell(row=o_n + 10, column=5).value = i_wb.cell(row=i_n, column=3).value  # Birthday
            o_wb.cell(row=o_n + 10, column=6).value = i_wb.cell(row=i_n, column=4).value  # Gender
            o_wb.cell(row=o_n + 10, column=7).value = i_wb.cell(row=i_n, column=9).value  # Relationship
            if i_wb.cell(row=i_n, column=8).value is 'E':
                o_wb.cell(row=o_n + 10, column=8).value = i_wb.cell(row=i_n, column=6).value  # Medical Tier

        if template is '7':  # Aetna CAT
            o_n += 1
            o_wb.cell(row=o_n + 4, column=1).value = i_wb.cell(row=i_n, column=2).value  # Last Name
            o_wb.cell(row=o_n + 4, column=2).value = i_wb.cell(row=i_n, column=1).value  # First Name
            o_wb.cell(row=o_n + 4, column=3).value = i_wb.cell(row=i_n, column=5).value  # Zip
            o_wb.cell(row=o_n + 4, column=4).value = i_wb.cell(row=i_n, column=3).value  # Birthday
            o_wb.cell(row=o_n + 4, column=5).value = i_wb.cell(row=i_n, column=4).value  # Gender
            o_wb.cell(row=o_n + 4, column=6).value = i_wb.cell(row=i_n, column=10).value  # Relationship Code

        if template is '8':  # Aetna Tier
            if i_wb.cell(row=i_n, column=8).value is 'E':
                o_n += 1
                o_wb.cell(row=o_n + 1, column=1).value = i_wb.cell(row=i_n, column=2).value  # Last Name
                o_wb.cell(row=o_n + 1, column=2).value = i_wb.cell(row=i_n, column=1).value  # First Name
                o_wb.cell(row=o_n + 1, column=3).value = i_wb.cell(row=i_n, column=5).value  # Zip
                o_wb.cell(row=o_n + 1, column=4).value = i_wb.cell(row=i_n, column=3).value  # Birthday
                o_wb.cell(row=o_n + 1, column=5).value = i_wb.cell(row=i_n, column=4).value  # Gender
                o_wb.cell(row=o_n + 1, column=6).value = i_wb.cell(row=i_n, column=6).value  # Medical Tier

        if template is '9':  # Anthem ACA
            o_n += 1
            o_wb.cell(row=o_n + 1, column=1).value = i_wb.cell(row=i_n, column=1).value  # First Name
            o_wb.cell(row=o_n + 1, column=2).value = i_wb.cell(row=i_n, column=2).value  # Last Name
            o_wb.cell(row=o_n + 1, column=3).value = i_wb.cell(row=i_n, column=4).value  # Gender
            o_wb.cell(row=o_n + 1, column=4).value = member_type[i_wb.cell(row=i_n, column=8).value]  # Member Type
            o_wb.cell(row=o_n + 1, column=5).value = i_wb.cell(row=i_n, column=3).value  # Birthday
            o_wb.cell(row=o_n + 1, column=6).value = i_wb.cell(row=i_n, column=5).value  # Zip
            if i_wb.cell(row=i_n, column=8).value is 'E':
                o_wb.cell(row=o_n + 1, column=7).value = "No"  # Cobra
                o_wb.cell(row=o_n + 1, column=8).value = tier_type[i_wb.cell(row=i_n, column=6).value]  # Tier Type

        i_n += 1

    o_file.save(new_dir)
    rename(file_loc, new_dir, template)


# Fills out remaining info on census file
def auto_fill(file_loc, include_child_count):
    file = load_workbook(filename=file_loc)
    wb = file.active
    n = 2
    child_count = 0
    while wb.cell(row=n, column=1).value is not None:
        # Ignore Waivers
        if wb.cell(row=n, column=6).value is 'W':
            n += 1
            continue

        # Auto-fill zip code
        if wb.cell(row=n, column=5).value is None and wb.cell(row=n, column=6).value is None:
            wb.cell(row=n, column=5).value = wb.cell(row=n - 1, column=5).value

        # If the person is only by themselves, then add in the info for single employers
        if wb.cell(row=n, column=6).value is ('E' or 'R' or 'C'):
            wb.cell(row=n, column=7).value = 'EE'
            wb.cell(row=n, column=8).value = 'E'
            wb.cell(row=n, column=9).value = 0
            wb.cell(row=n, column=10).value = 1

        # Get relationship, E/ES/EC/F
        if wb.cell(row=n, column=6).value is not 'E' and wb.cell(row=n, column=7).value is None:
            # Is this person the provider?
            if wb.cell(row=n, column=6).value is not None:
                wb.cell(row=n, column=7).value = \
                    str(wb.cell(row=n, column=6).value) if wb.cell(row=n, column=6).value not in 'CR' else 'EE'
                # if wb.cell(row=n, column=6).value is 'R' or 'C':
                #     wb.cell(row=n, column=7).value = 'E'
                wb.cell(row=n, column=8).value = 'E'
                wb.cell(row=n, column=9).value = 0
                wb.cell(row=n, column=10).value = 1
                child_count = 0
            # Is this person a family member?
            else:
                wb.cell(row=n, column=6).value = str(wb.cell(row=n - 1, column=6).value)
                wb.cell(row=n, column=7).value = str(wb.cell(row=n - 1, column=7).value)

                # Is the person above an employee?
                if wb.cell(row=n - 1, column=8).value is 'E':
                    if 'F' in wb.cell(row=n, column=6).value or 'ES' in wb.cell(row=n, column=6).value:
                        wb.cell(row=n, column=8).value = 'S'
                        wb.cell(row=n, column=9).value = 1
                        wb.cell(row=n, column=10).value = 0
                    if 'EC' in wb.cell(row=n, column=6).value:
                        wb.cell(row=n, column=8).value = 'D'
                        wb.cell(row=n, column=9).value = 2
                        wb.cell(row=n, column=10).value = 0

                # Is the person above a spouse with a family?
                if wb.cell(row=n - 1, column=8).value is 'S' and 'F' in wb.cell(row=n, column=6).value:
                    wb.cell(row=n, column=8).value = 'D'
                    wb.cell(row=n, column=9).value = 2
                    wb.cell(row=n, column=10).value = 0

                # Is the person above a child?
                if wb.cell(row=n - 1, column=8).value is 'D':
                    # Had to separate the if statements cause Python is doing something weird with it
                    if 'F' in wb.cell(row=n, column=6).value or 'EC' in wb.cell(row=n, column=6).value:
                        wb.cell(row=n, column=8).value = 'D'
                        wb.cell(row=n, column=9).value = 2
                        wb.cell(row=n, column=10).value = 0

        wb.cell(row=n, column=11).value = None
        if wb.cell(row=n, column=8).value is 'D' and include_child_count:
            child_count += 1
            is_spouse = 0
            if wb.cell(row=n - child_count, column=8).value is 'S':
                is_spouse = 1
            wb.cell(row=n - child_count - is_spouse, column=11).value = child_count
        else:
            child_count = 0

        if wb.cell(row=n, column=3).value is not None:
            try:
                wb.cell(row=n, column=12).value = datetime.strptime(str(wb.cell(row=n, column=3).value),
                                                                    '%m/%d/%Y').strftime('%m/%y')
            except(ValueError, Exception):
                wb.cell(row=n, column=3).value = datetime.strptime(str(wb.cell(row=n, column=3).value),
                                                                   '%Y-%m-%d 00:00:00').strftime('%m/%d/%Y')
                wb.cell(row=n, column=12).value = datetime.strptime(str(wb.cell(row=n, column=3).value),
                                                                    '%m/%d/%Y').strftime('%m/%y')

        if get_age(wb.cell(row=n, column=3).value) >= 65:
            for col in range(1, 13):
                wb.cell(row=n, column=col).fill = PatternFill(fill_type='solid', start_color='B8FAFA',
                                                              end_color='B8FAFA')
        if get_age(wb.cell(row=n, column=3).value) >= 26 and wb.cell(row=n, column=8).value is "D":
            for col in range(1, 13):
                wb.cell(row=n, column=col).fill = PatternFill(fill_type='solid', start_color='F2F277',
                                                              end_color='F2F277')
        n += 1

    file.save(filename=file_loc)


def create_ft_census(file, wait):
    tnf_file = load_workbook(file)
    tnf_wb = tnf_file.active

    # Move the template over to the correct location
    new_loc = file.replace(os.path.basename(file), '')  # Get target directory
    copy2('templates/Completed FT Census.xlsx', new_loc)  # Copy the new file into the target directory
    new_dir = os.path.join(new_loc, os.path.basename('templates/Completed FT Census.xlsx'))
    ft_file = load_workbook(new_dir)  # Load the copied file
    ft_wb = ft_file.active

    tnf_row = 8
    # Look into if the .strip function works
    ein = str(tnf_wb.cell(row=5, column=1).value)[len("Company EIN: "):]

    while tnf_wb.cell(row=tnf_row, column=1).value is not None:
        # Place info into the sheet
        ft_wb.cell(row=tnf_row + 16, column=1).value = ein
        # Check if social security numbers don't have a dash in them already, if not add them
        if "-" not in str(tnf_wb.cell(row=tnf_row, column=9).value):
            ssn = str(tnf_wb.cell(row=tnf_row, column=9).value)[:3] + "-" + str(
                tnf_wb.cell(row=tnf_row, column=9).value)[3:]
            ssn = ssn[:6] + "-" + ssn[6:]
        else:
            ssn = str(tnf_wb.cell(row=tnf_row, column=9).value)
        ft_wb.cell(row=tnf_row + 16, column=2).value = ssn  # Employee SSN
        ft_wb.cell(row=tnf_row + 16, column=3).value = ssn  # Member SSN
        ft_wb.cell(row=tnf_row + 16, column=4).value = tnf_wb.cell(row=tnf_row, column=2).value  # First Name
        ft_wb.cell(row=tnf_row + 16, column=6).value = tnf_wb.cell(row=tnf_row, column=1).value  # Last Name
        ft_wb.cell(row=tnf_row + 16, column=11).value = tnf_wb.cell(row=tnf_row, column=4).value  # Address 1
        ft_wb.cell(row=tnf_row + 16, column=12).value = tnf_wb.cell(row=tnf_row, column=5).value  # Address 2
        ft_wb.cell(row=tnf_row + 16, column=13).value = tnf_wb.cell(row=tnf_row, column=6).value  # City
        ft_wb.cell(row=tnf_row + 16, column=14).value = tnf_wb.cell(row=tnf_row, column=7).value  # State
        ft_wb.cell(row=tnf_row + 16, column=15).value = tnf_wb.cell(row=tnf_row, column=8).value  # Zip

        is_empty = True
        for col in range(0, 13):
            if 'x' in str(tnf_wb.cell(row=tnf_row, column=10 + col).value).lower():
                is_empty = False
                break

        # Monthly coverage
        if not is_empty:
            month_count = 0
            for col in range(0, 12):
                # Offer of coverage and enrollment, only if non-yearly isn't selected
                if 'x' not in str(tnf_wb.cell(row=tnf_row, column=10).value).lower():
                    if 'x' in str(tnf_wb.cell(row=tnf_row, column=11 + col).value).lower():
                        ft_wb.cell(row=tnf_row + 16, column=27 + (col * 3)).value = "1A"
                        ft_wb.cell(row=tnf_row + 16, column=29 + (col * 3)).value = "2C"
                        if month_count is 0:
                            month_count = col + 1
                    else:
                        ft_wb.cell(row=tnf_row + 16, column=27 + (col * 3)).value = "1H"
                        ft_wb.cell(row=tnf_row + 16, column=29 + (col * 3)).value = "2A"

            for col in range(0, 12):
                if month_count is 0:
                    ft_wb.cell(row=tnf_row + 16, column=27 + (col * 3)).value = None
                    ft_wb.cell(row=tnf_row + 16, column=29 + (col * 3)).value = None
                # 11 - 2 = 10 + 1
                elif month_count - wait <= col + 1 < month_count:
                    ft_wb.cell(row=tnf_row + 16, column=29 + (col * 3)).value = "2D"

            # Check if employee had no coverage in the middle of the year. May revise later
            for col in range(0, 11):
                if ft_wb.cell(row=tnf_row + 16, column=29 + (wait * 3) + (col * 3)).value is ("2C" or None) and \
                        ft_wb.cell(row=tnf_row + 16, column=29 + (col * 3)).value is "2A":
                    ft_wb.cell(row=tnf_row + 16, column=29 + (col * 3)).value = "2D"

            # Yearly coverage, if applicable
            if 'x' in str(tnf_wb.cell(row=tnf_row, column=10).value).lower():
                ft_wb.cell(row=tnf_row + 16, column=24).value = "1A"
                ft_wb.cell(row=tnf_row + 16, column=26).value = "2C"
            elif month_count is 0:
                ft_wb.cell(row=tnf_row + 16, column=24).value = "1A"
                ft_wb.cell(row=tnf_row + 16, column=26).value = None

            # Set format to 2 digits and calculate starting month
            # ft_wb.cell(row=tnf_row + 16, column=23).number_format = "00"
            ft_wb.cell(row=tnf_row + 16, column=23).value = month_count if month_count is not 0 else None
            if month_count is 0:
                ft_wb.cell(row=tnf_row + 16, column=23).value = 1

        tnf_row += 1

    ft_file.save(new_dir)
    xlsx_to_csv(new_dir)


# Auto-fill the ICHRA quoting census, using the Marketplace API to retrieve quotes for each member
# Key: LRnyaUBUc97sbfT9FSy5US2UtbdiryA5
def create_ichra_plans(file, year, metal_levels):
    # Load the selected file
    ichra_file = load_workbook(file)
    ichra_wb = ichra_file.active
    ichra_row = 2
    # API Key and the url for searching plans
    key = "LRnyaUBUc97sbfT9FSy5US2UtbdiryA5"
    plan_url = "https://marketplace.api.healthcare.gov/api/v1/plans/search?apikey=" + key + "&year=" + str(year)
    # Set up the header and body for the request
    hdr = {'Content-Type': 'application/json'}
    marketplace_body = {
        "household": {
            "people": [{
                "dob": "",
                "has_mec": False,
                "uses_tobacco": False,
                "utilization_level": "Low"
            }]
        },
        "filter": {
            "metal_levels": metal_levels
        },
        "market": "Individual",
        "place": {
            "countyfips": "",
            "state": "",
            "zipcode": ""
        },
        "limit": 0,
        "offset": 0,
        "order": "asc",
        "year": int(year)
    }
    # Create a plan list to keep track of plan names
    plan_list = []

    # Remove any previous plans from the spread
    ichra_wb.delete_cols(10, 200)

    # Go through every applicable row of the sheet, skipping over dependents
    while ichra_wb.cell(row=ichra_row, column=1).value is not None:
        if ichra_wb.cell(row=ichra_row, column=7).value is not None:
            # Retrieve the member county by sending a GET request and placing the results in the marketplace body
            try:
                zipcode = str(ichra_wb.cell(row=ichra_row, column=8).value)
                county_url = \
                    "https://marketplace.api.healthcare.gov/api/v1/counties/by/zip/" + zipcode + "?apikey=" + key
                county_query = requests.get(county_url, headers=hdr).json()["counties"]
                found_county = False
                if ichra_wb.cell(row=ichra_row, column=9).value is not None:
                    for county in county_query:
                        if ichra_wb.cell(row=ichra_row, column=9).value in county["name"]:
                            county_query = county
                            found_county = True
                            break
                if not found_county:
                    county_query = county_query[0]
            except(IndexError, Exception):
                print(ichra_wb.cell(row=ichra_row, column=1).value + " " + ichra_wb.cell(row=ichra_row, column=2).value
                      + " - Plans retrieved: 0 - County Not Found")
                ichra_row += 1
                continue
            # If counties have been retrieved, add the county name to the row and complete the remaining actions
            if ichra_wb.cell(row=ichra_row, column=9).value is None:
                ichra_wb.cell(row=ichra_row, column=9).value = str(county_query["name"]).strip(" County")
            marketplace_body["place"]["countyfips"] = county_query["fips"]
            marketplace_body["place"]["state"] = county_query["state"]
            marketplace_body["place"]["zipcode"] = county_query["zipcode"]
            # Add DoB of main members, formatting it correctly
            marketplace_body["household"]["people"][0]["dob"] = \
                ichra_wb.cell(row=ichra_row, column=4).value.strftime("%Y-%m-%d")
            # Remove any previous dependents and check for dependents of the current member
            marketplace_body["household"]["people"] = marketplace_body["household"]["people"][0:1]
            dep_i = 1
            while ichra_wb.cell(row=ichra_row + dep_i, column=7).value is None and \
                    ichra_wb.cell(row=ichra_row + dep_i, column=1).value is not None:
                marketplace_body["household"]["people"].append({
                    "dob": ichra_wb.cell(row=ichra_row + dep_i, column=4).value.strftime("%Y-%m-%d"),
                    "has_mec": False,
                    "uses_tobacco": False,
                    "utilization_level": "Low"
                })
                dep_i += 1
            # Create the a list of plans for this row
            row_plan = generate_plan_dict(plan_url, hdr, marketplace_body)

            for plan in row_plan:
                # If the plan is not in the list, set up a column for it
                if plan[1] not in plan_list:
                    plan_list.append(plan[1])
                    # Go to the correct column
                    col = 9 + len(plan_list)
                    # Change column width to fit the new plan a bit easier
                    ichra_wb.column_dimensions[get_column_letter(col)].width = 20
                    # Get the correct Ded amount and create the title
                    mult = 1 if len(marketplace_body["household"]["people"]) > 1 else 2
                    title = ichra_wb.cell(row=1, column=col)
                    title.value = re.sub(r'\([^()]*\)', '', plan[1]) + "\n(" + str(int(plan[3] * mult / 2)) + "/" + str(
                        plan[3] * mult) + " Ded)"
                    title.font = Font(bold=True)
                    title.alignment = Alignment(wrap_text=True)
                # If the plan is already in the list, go to that plan column
                else:
                    col = 10 + plan_list.index(plan[1])
                # Place the premium into the correct row and column
                # ichra_wb.cell(row=ichra_row, column=col).value = plan[2]
                premium = ichra_wb.cell(row=ichra_row, column=col)
                premium.value = plan[2]
                premium.number_format = '$#.00'
                # print(str(ichra_row) + " : " + str(col) + " : " + str(plan[2]))

            print(ichra_wb.cell(row=ichra_row, column=1).value + " " + ichra_wb.cell(row=ichra_row, column=2).value +
                  " - Plans retrieved: " + str(len(row_plan)) + " - " + county_query["name"])

        ichra_row += 1

    # Once the loop has been completed, add the totals
    total = ichra_wb.cell(row=ichra_row + 1, column=9)
    total.value = "Totals"
    total.font = Font(bold=True)
    for t_col in range(10, 10 + len(plan_list)):
        ichra_wb.cell(row=ichra_row + 1, column=t_col).font = Font(bold=True)
        ichra_wb.cell(row=ichra_row + 1, column=t_col).number_format = "$0.00"
        ichra_wb.cell(row=ichra_row + 1, column=t_col).value = \
            "=SUM(" + get_column_letter(t_col) + "2:" + get_column_letter(t_col) + str(ichra_row - 1) + ")"

    ichra_file.save(filename=file)


def generate_plan_dict(url, header, body):
    # Set up a plan array, going through the full query and retrieve the name, premium amount, and deductibles
    plan_array = []
    plan_loop = True
    # Retrieve a dictionary of plans
    while plan_loop:
        plan_query = requests.post(url, headers=header, json=body).json()["plans"]
        for i in range(0, len(plan_query)):
            plan_array.append([plan_query[i]["issuer"]["name"], plan_query[i]["name"], plan_query[i]["premium"],
                               plan_query[i]["deductibles"][0]["amount"]])
            # print(plan_array[i + body["offset"]])
        # If the query is 10, there may be more plans to retrieve
        if len(plan_query) == 10:
            body["offset"] = body["offset"] + 10
        else:
            plan_loop = False
            body["offset"] = 0

    return plan_array
